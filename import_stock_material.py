"""
Import stock material data from 'Total Stock List.xlsx' into the stockMaterial table.

Usage:
    cd panamera_backend
    python manage.py shell < import_stock_material.py

    OR run directly:
    python import_stock_material.py
"""

import os
import sys
import django

# Setup Django if running standalone
if __name__ == "__main__":
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), "panamera_backend"))
    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "panamera_backend.settings")
    django.setup()

import openpyxl
from django.db import connection


EXCEL_PATH = os.path.join(os.path.dirname(__file__), "Total Stock List.xlsx")
BATCH_SIZE = 500


def import_stock_materials():
    """Read Excel and bulk-insert into stockMaterial table."""
    print(f"Loading workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb.active

    # Validate headers
    headers = [cell.value for cell in ws[1]]
    expected = ["Stock Code", "Description", "Unit"]
    if headers[:3] != expected:
        print(f"ERROR: Expected headers {expected}, got {headers[:3]}")
        return

    rows_to_insert = []
    skipped = 0
    seen_codes = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        stock_code = str(row[0] or "").strip()
        description = str(row[1] or "").strip()
        unit = str(row[2] or "").strip()

        # Skip rows with empty stock code
        if not stock_code:
            skipped += 1
            continue

        # Skip duplicate stock codes within the Excel file
        if stock_code in seen_codes:
            skipped += 1
            continue
        seen_codes.add(stock_code)

        rows_to_insert.append((stock_code, description, unit))

    wb.close()
    print(f"Parsed {len(rows_to_insert)} unique rows ({skipped} skipped)")

    if not rows_to_insert:
        print("No data to insert.")
        return

    # Bulk insert in batches using ON CONFLICT to handle any pre-existing data
    insert_sql = """
        INSERT INTO public."stockMaterial" ("stockCode", description, unit, "createdAt", "updatedAt", "isDeleted")
        VALUES (%s, %s, %s, NOW(), NOW(), 0)
        ON CONFLICT ("stockCode") DO UPDATE
            SET description = EXCLUDED.description,
                unit = EXCLUDED.unit,
                "updatedAt" = NOW()
    """

    inserted = 0
    with connection.cursor() as cursor:
        for i in range(0, len(rows_to_insert), BATCH_SIZE):
            batch = rows_to_insert[i : i + BATCH_SIZE]
            for stock_code, description, unit in batch:
                cursor.execute(insert_sql, [stock_code, description, unit])
            inserted += len(batch)
            print(f"  Processed {inserted}/{len(rows_to_insert)} rows...")

    print(f"\nDone! {inserted} rows inserted/updated in stockMaterial table.")

    # Verify count
    with connection.cursor() as cursor:
        cursor.execute('SELECT COUNT(*) FROM public."stockMaterial" WHERE "isDeleted" = 0')
        count = cursor.fetchone()[0]
    print(f"Total active rows in stockMaterial: {count}")


if __name__ == "__main__":
    import_stock_materials()
